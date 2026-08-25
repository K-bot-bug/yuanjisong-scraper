"""
对python-freshman.xlsx进行大类分类
"""
import openpyxl, sys, re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\23188\Desktop\test\爬取猿急送\数据\python-freshman.xlsx')
ws = wb.active


def categorize(title, desc):
    text = f'{title} {desc}'.lower()

    # ========== 爬虫/数据采集（最优先，避免被其他规则误捕获）==========
    if any(kw in text for kw in ['爬虫', '爬取', '抓取', '采集', '反爬', '绕过验证码',
                                   '数据爬取', '批量爬取', '网页抓取', '数据抓取',
                                   'web scraping', '爬虫脚本']):
        return '爬虫-数据采集'

    # ========== 小程序/移动端（优先，避免被Python/后端规则误捕获）==========
    if any(kw in text for kw in ['小程序', '微信小程序', 'uni-app', 'uniapp', 'taro',
                                   'mp-weixin', '微信公众号', '微信开发', '小程序商城',
                                   'app开发', 'android开发', 'ios开发', 'flutter',
                                   'react-native', '鸿蒙', 'android客户端',
                                   'app上架', 'app测试', 'app二开', 'app加固',
                                   '混淆加壳', '移动应用']):
        return '小程序/移动端'

    # ========== 前端/Web页面 ==========
    if any(kw in text for kw in ['vue', 'react', 'angular', 'h5', 'html', 'css',
                                   '前端开发', '前端框架', '静态页面', 'ui美化',
                                   'ui设计', '网页制作', '网站改版', '页面开发',
                                   '响应式', 'bootstrap', 'element-ui', 'antd',
                                   'tailwind', 'javascript开发', 'js页面',
                                   'web页面', '网页开发', 'pc端页面',
                                   '管理后台', '后台页面', 'webview', 'h5页面',
                                   '网页设计', '界面开发', '前端二开', '网页bug']):
        return '前端/Web页面'

    # ========== Python脚本 ==========
    if any(kw in text for kw in ['python', 'pandas', 'pyqt', 'tkinter', 'pygame',
                                   'opencv', 'pil', ' Pillow ', 'image', '图片处理',
                                   '批量处理', '数据清洗', '自动重命名',
                                   '表格处理', 'excel', 'csv', 'xlsx', '数据导入导出',
                                   '自动化脚本', '定时任务', '批处理', '数据转换',
                                   'json处理', 'xml处理', 'email自动化', '文件处理',
                                   '数据可视化', 'matplotlib', '数据导出', '批量操作',
                                   '办公自动化']):
        return 'Python脚本'

    # ========== 后端/接口/部署 ==========
    if any(kw in text for kw in ['api调试', 'api对接', '接口调试', '接口开发',
                                   '接口联调', '签名验证', '鉴权', 'webhook',
                                   '后端开发', '后端框架', 'java开发', 'spring',
                                   'fastapi', 'django', 'flask', 'node.js',
                                   'php开发', 'go开发', 'restful', 'graphql',
                                   '数据库', 'mysql', 'mongodb', 'redis',
                                   '服务器部署', 'docker', 'nginx', '部署上线',
                                   '运维', '环境搭建', '数据库迁移', '服务器迁移',
                                   '后端二开', '二次开发', 'php后端',
                                   'erp', 'oa系统', 'cms开发', '后台管理',
                                   '权限管理', '工作流', '低代码', '微服务',
                                   '消息队列', 'kafka', 'rabbitmq',
                                   # 补充更多关键词
                                   '接口对接', 'api接入', 'api集成', '接口联入',
                                   'java问题', 'java开发', 'java修复',
                                   'php开发', 'php修改', 'php对接',
                                   'c#开发', 'c#后端', 'c#修复', 'csharp',
                                   '服务器迁移', '服务器部署', '服务器维护',
                                   '数据库迁移', '数据库开发', '数据库修复',
                                   '短信接口', '支付接口', '物流接口',
                                   'api接入', '接口问题', '接口修复',
                                   '网站后端', '后台开发', '后台修复',
                                   '电商开发', '商城开发', '商城系统',
                                   'oa开发', 'erp开发', '政务系统',
                                   '网站维护', '网站修复', '网站bug',
                                   '服务器安全', '安全修复', '病毒排查']):
        return '后端/接口/部署'

    # ========== 测试/质检 ==========
    if any(kw in text for kw in ['软件测试', '自动化测试', 'app测试', '性能测试',
                                   '安全测试', '兼容性测试', '压力测试',
                                   'bug修复', '质量检查', '测试脚本',
                                   '单元测试', '集成测试', '接口测试',
                                   '测试工具', 'web自动化测试', '测试开发',
                                   '功能测试', '回归测试']):
        return '测试/质检'

    # ========== AI/智能体 ==========
    if any(kw in text for kw in ['ai应用', '大模型', 'llm', 'chatgpt', '智能体',
                                   'comfyui', 'stable diffusion', 'midjourney',
                                   'rag', '向量数据库', 'embedding', 'nlp',
                                   '语音识别', '图像识别', '机器学习',
                                   '深度学习', '神经网络', 'ocr识别',
                                   '人工智能', 'ai开发', 'pytorch', 'tensorflow',
                                   'gcn', 'gpt', 'transformer', 'bert',
                                   '模型训练', '模型微调', '微调', '推理',
                                   '视觉模型', 'nlu', 'nmt']):
        return 'AI/智能体'

    # ========== 嵌入式/IoT/硬件 ==========
    if any(kw in text for kw in ['stm32', '嵌入式', 'iot', 'esp32', '树莓派',
                                   '单片机', '硬件开发', '传感器', '物联网',
                                   'fpga', '嵌入式开发', '电路板', '固件开发',
                                   '驱动开发', 'arm开发', 'bluetooth', '蓝牙',
                                   'arduino', 'raspberry', 'mcu', 'gpio',
                                   '电机控制', '嵌入式linux']):
        return '嵌入式/IoT/硬件'

    # ========== 工具/其他脚本 ==========
    if any(kw in text for kw in ['桌面应用', 'electron', 'qt开发', 'windows应用',
                                   '浏览器插件', '油猴', 'tampermonkey', '用户脚本',
                                   'chrome扩展', '小工具', '工具开发',
                                   '二维码', '图片处理工具', '视频剪辑']):
        return '工具/其他脚本'

    # ========== 通用项目 ==========
    return '通用项目'


rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    if len(vals) >= 9:
        idx, title, status, hours, price, apply_count, publisher, desc, link = vals[:9]
    else:
        continue
    cat = categorize(title, desc)
    rows.append((idx, title, status, hours, price, apply_count, publisher, desc, link, cat))

cat_counts = Counter(r[9] for r in rows)

out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = '汇总'

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
CATEG_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

headers = ['序号', '项目标题', '项目状态', '工时', '预算', '投递人数', '发布者', '项目描述', '项目链接', '分类']
for col, h in enumerate(headers, 1):
    c = out_ws.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN_BORDER

for i, r in enumerate(rows, 1):
    row_idx = i + 1
    for col, v in enumerate(r, 1):
        c = out_ws.cell(row=row_idx, column=col, value=v)
        c.border = THIN_BORDER
        if col == 1:
            c.alignment = Alignment(horizontal='center')
    out_ws.cell(row=row_idx, column=10).fill = CATEG_FILL

out_ws.column_dimensions['A'].width = 6
out_ws.column_dimensions['B'].width = 45
out_ws.column_dimensions['C'].width = 18
out_ws.column_dimensions['D'].width = 10
out_ws.column_dimensions['E'].width = 12
out_ws.column_dimensions['F'].width = 12
out_ws.column_dimensions['G'].width = 18
out_ws.column_dimensions['H'].width = 80
out_ws.column_dimensions['I'].width = 45
out_ws.column_dimensions['J'].width = 18

CAT_COLORS = {
    '爬虫-数据采集':     ('E3F2FD', '1565C0'),
    'Python脚本':       ('F3E5F5', '7B1FA2'),
    '小程序/移动端':    ('FFF3E0', 'E65100'),
    '前端/Web页面':     ('E8F5E9', '2E7D32'),
    '后端/接口/部署':   ('FCE4EC', 'C62828'),
    '测试/质检':        ('E0F7FA', '00838F'),
    'AI/智能体':        ('F1F8E9', '558B2F'),
    '嵌入式/IoT/硬件':  ('EDE7F6', '4527A0'),
    '工具/其他脚本':    ('FFF9C4', 'F57F17'),
    '通用项目':         ('EEEEEE', '616161'),
}

for cat in sorted(cat_counts.keys(), key=lambda x: -cat_counts[x]):
    cat_rows = [r for r in rows if r[9] == cat]
    sheet_name = cat.replace('/', '-').replace(' ', '_')[:31]
    new_ws = out_wb.create_sheet(title=sheet_name)

    fill_color, _ = CAT_COLORS.get(cat, ('EEEEEE', '333333'))
    cat_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    for col, h in enumerate(headers, 1):
        c = new_ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = THIN_BORDER

    for i, r in enumerate(cat_rows, 1):
        row_idx = i + 1
        for col, v in enumerate(r, 1):
            c = new_ws.cell(row=row_idx, column=col, value=v)
            c.border = THIN_BORDER
            if col == 1:
                c.alignment = Alignment(horizontal='center')
            if col == 10:
                c.fill = cat_fill

    new_ws.column_dimensions['A'].width = 6
    new_ws.column_dimensions['B'].width = 45
    new_ws.column_dimensions['C'].width = 18
    new_ws.column_dimensions['D'].width = 10
    new_ws.column_dimensions['E'].width = 12
    new_ws.column_dimensions['F'].width = 12
    new_ws.column_dimensions['G'].width = 18
    new_ws.column_dimensions['H'].width = 80
    new_ws.column_dimensions['I'].width = 45
    new_ws.column_dimensions['J'].width = 18

out_path = r'C:\Users\23188\Desktop\test\爬取猿急送\数据\python-projects-v2.xlsx'
out_wb.save(out_path)

print('=' * 60)
print('分类统计结果')
print('=' * 60)
for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
    print(f'  {cat}: {count} 条')
print(f'\n总计: {len(rows)} 条')
print(f'\n已保存: {out_path}')
