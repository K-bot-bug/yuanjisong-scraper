"""
筛选500元以内、适合大学生的项目，导出到student-friendly.xlsx
"""
import openpyxl, sys, re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
sys.stdout.reconfigure(encoding='utf-8')

# 读取原始数据
import os, sys
sys.stdout.reconfigure(encoding='utf-8')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "数据")
DEFAULT_INPUT = os.path.join(DATA_DIR, "yuanjisong_jobs.xlsx")
INPUT_FILE = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_INPUT

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

# 排除关键词（高难度/违规/敏感）
exclude_kw = [
    '架构', '高级', '专家', '深入', '底层', '内核', '编译原理', '逆向工程', '风控',
    '算法', '深度学习', '区块链', '量化', '挖矿', '渗透测试', '漏洞挖掘', '密码学',
    '安全研究', '反作弊', '反欺诈', '数字取证', '威胁情报', '性能优化', '架构设计',
    '反调试', 'root分析', 'so文件', '脱壳', 'jni', 'ndk', '驱动开发',
    'unity3d', '虚幻引擎', 'cocos', '游戏服务器', 'mmo',
    '接码平台', '翻墙', 'vpn代理', '绕过', '破解', '灰色', '套利', '刷单',
    '刷量', '黑产', '洗钱', '赌博', '彩票', '博彩', '棋牌', '网赌',
    '代写论文', '代做毕设', '学术不端',
    'stm32', '嵌入式', 'esp32', '树莓派', '单片机', 'iot', 'arduino',
    '传感器', '物联网', 'fpga', '电路板', '固件', 'arm开发',
    'app加固', '混淆', '加壳', 'so开发', '视频防嗅探',
    '游戏', 'unity', 'game', '网游', '手游',
    '相机', '标定', '镜头', '光学',
]

suitable = []
for row in ws.iter_rows(min_row=2, values_only=True):
    idx, title, status, hours, price, apply_count, publisher, desc = row[:8]
    p_match = re.search(r'(\d+)', str(price))
    p = int(p_match.group(1)) if p_match else 0
    # 只保留500元以下、非驻场的项目
    if p > 500 or '驻场' in str(status or ''):
        continue
    text = f'{title} {desc}'
    text_lower = text.lower()
    # 排除含敏感词的项目
    if any(kw.lower() in text_lower for kw in exclude_kw):
        continue
    # 生成链接
    link = f'https://www.yuanjisong.com/job/{idx}'
    suitable.append((idx, title, status, hours, price, apply_count, publisher, desc, link))

# 按预算排序
suitable.sort(key=lambda x: int(re.search(r'\d+', x[4]).group()) if re.search(r'\d+', x[4]) else 999)

# 导出Excel
new_wb = openpyxl.Workbook()
out_ws = new_wb.active
out_ws.title = '大学生适合项目'

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

headers = ['序号', '项目标题', '项目状态', '工时', '预算', '投递人数', '发布者', '项目描述', '项目链接']
for col, h in enumerate(headers, 1):
    c = out_ws.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = THIN_BORDER

for i, r in enumerate(suitable, 1):
    row_idx = i + 1
    vals = [i, r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]]
    for col, v in enumerate(vals, 1):
        c = out_ws.cell(row=row_idx, column=col, value=v)
        c.border = THIN_BORDER
        if col == 1:
            c.alignment = Alignment(horizontal='center')

out_ws.column_dimensions['A'].width = 6
out_ws.column_dimensions['B'].width = 45
out_ws.column_dimensions['C'].width = 18
out_ws.column_dimensions['D'].width = 10
out_ws.column_dimensions['E'].width = 12
out_ws.column_dimensions['F'].width = 12
out_ws.column_dimensions['G'].width = 18
out_ws.column_dimensions['H'].width = 80
out_ws.column_dimensions['I'].width = 45
for r in range(2, len(suitable) + 2):
    out_ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical='top')

out_path = r'C:\Users\23188\Desktop\test\爬取猿急送\数据\student-friendly.xlsx'
new_wb.save(out_path)

print(f'筛选条件: 预算<=500元 + 非驻场 + 排除{len(exclude_kw)}个高风险/高难度词')
print(f'适合大学生项目: {len(suitable)} 条')
print(f'已保存: {out_path}')
print()
print('前30条预览:')
for r in suitable[:30]:
    print(f'  {r[0]:>4}. {r[1][:50]:<50} | {r[2]:<14} | {r[3]:<8} | {r[4]:<10} | {r[6]}')
    print(f'      {r[8]}')
    print()
