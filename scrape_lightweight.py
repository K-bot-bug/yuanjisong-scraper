"""
猿急送爬虫 - curl_cffi 轻量版
==============================
优势：
- curl_cffi impersonate='chrome131' 完整模拟 Chrome TLS/HTTP2 指纹
- 无需启动浏览器进程，速度快、资源占用低
- 适用于纯 SSR 站点（猿急送无 JS 动态渲染）

使用方式：
  python scrape_lightweight.py [选项]

选项：
  --max-pages N     最大页数（默认 300）
  --output-dir DIR  输出目录（默认 ./数据）
  --include         不排除任何关键词（调试用）
"""
import asyncio
import json
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from curl_cffi import requests as cffi_requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from loguru import logger

from proxy_pool import get_pool, get_proxy_url, init_pool

sys.stdout.reconfigure(encoding='utf-8')
logger.add("scrape_lightweight.log", rotation="10 MB", encoding="utf-8")

# ========== 配置 ==========
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "数据"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

BASE_URL = "https://www.yuanjisong.com/job/allcity"
COOKIE_FILE = SCRIPT_DIR / ".cookies.json"
LAST_PAGE_FILE = SCRIPT_DIR / ".last_page.txt"

# 排除关键词
EXCLUDE_KW = [
    '架构', '高级', '专家', '深入', '底层', '内核', '编译原理', '逆向工程', '风控',
    '算法', '深度学习', '区块链', '量化', '挖矿', '渗透测试', '漏洞挖掘', '密码学',
    '安全研究', '反作弊', '反欺诈', '数字取证', '威胁情报', '性能优化',
    '接码平台', '翻墙', 'vpn代理', '绕过', '破解', '灰色', '套利', '刷单',
    '刷量', '黑产', '洗钱', '赌博', '彩票', '博彩', '棋牌', '网赌',
    '代写论文', '代做毕设', '学术不端',
    'stm32', '嵌入式', 'esp32', '树莓派', '单片机', 'iot', 'arduino',
    '传感器', '物联网', 'fpga', '电路板', '固件', 'arm开发',
    'app加固', '混淆', '加壳', 'so开发', '视频防嗅探',
    '游戏', 'unity', 'game', '网游', '手游',
    '相机', '标定', '镜头', '光学', '驻场',
]

# Excel 样式
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)


# ========== Cookie 持久化 ==========
def save_cookies(session: cffi_requests.Session, path: Path):
    """将当前 session 的 cookie 保存到文件（保留 domain/path/expires）"""
    cookies = {}
    for c in session.cookies.jar:
        cookies[c.name] = {
            'value': c.value,
            'domain': c.domain,
            'path': c.path,
            'secure': c.secure,
            'expires': c.expires,
        }
    try:
        path.write_text(json.dumps(cookies, ensure_ascii=False), encoding='utf-8')
        logger.debug(f"Cookie 已保存: {len(cookies)} 个")
    except Exception:
        pass


def load_cookies(session: cffi_requests.Session, path: Path) -> bool:
    """从文件加载 cookie 到 session"""
    if not path.exists():
        return False
    try:
        cookies = json.loads(path.read_text(encoding='utf-8'))
        for name, info in cookies.items():
            session.cookies.set(
                name, info['value'],
                domain=info.get('domain'),
                path=info.get('path', '/'),
            )
        logger.debug(f"Cookie 已加载: {len(cookies)} 个")
        return True
    except Exception:
        return False


def save_last_page(page_num: int, path: Path):
    """记录最后成功抓取的页码（断点续爬用）"""
    try:
        path.write_text(str(page_num), encoding='utf-8')
    except Exception:
        pass


def load_last_page(path: Path) -> int:
    """加载上次抓取的页码"""
    if not path.exists():
        return 1
    try:
        return int(path.read_text(encoding='utf-8').strip())
    except Exception:
        return 1


def parse_page(html: str) -> list[dict]:
    """从 HTML 中提取项目列表（使用 BeautifulSoup 处理嵌套 div）"""
    soup = BeautifulSoup(html, 'html.parser')
    jobs = []
    for card in soup.find_all('div', class_='job_card'):
        # 必须有可投递按钮才收录
        btn = card.find('a', class_='job_btn_apply')
        if not btn or 'disabled' in btn.get('class', []):
            continue

        def g(sel):
            el = card.select_one(sel)
            return el.get_text(strip=True) if el else ''

        title_link = card.find('a', class_='job_card_title_link')
        if not title_link:
            continue
        link = title_link.get('href', '')
        if not link.startswith('http'):
            link = 'https://www.yuanjisong.com' + link
        job_id = link.split('/')[-1]
        if not job_id:
            continue

        status = g('.job_tag_type')
        hours_raw = g('.job_tag_hours')
        hours = re.sub(r'<[^>]+>', '', hours_raw).replace('工时：', '').strip()
        price = g('.job_card_price')
        desc_raw = g('.job_card_desc')
        description = desc_raw.replace('描述：', '').strip()
        apply_count = g('.job_card_postnum')
        publisher = g('.job_card_publisher_name')

        jobs.append({
            'id': job_id,
            'title': g('.job_card_title'),
            'status': status,
            'hours': hours,
            'price': price,
            'description': description,
            'apply_count': apply_count,
            'publisher': publisher,
            'link': link,
        })
    return jobs


def _clean_for_excel(text) -> str:
    """移除 Excel 单元格不支持的非法字符（换行符、控制字符等）"""
    if not text:
        return ''
    # 替换控制字符为空格（保留正常换行用于 wrap_text）
    import re as _re
    return _re.sub(r'[\x00-\x08\x0a\x0b\x0c\x0d\x0e-\x1f]', '', str(text))


def save_excel(jobs: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = '可投递项目'

    headers = ['序号', '项目标题', '项目状态', '工时', '预算', '投递人数', '发布者', '项目描述', '项目链接']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = THIN_BORDER

    for i, j in enumerate(jobs, 1):
        row_idx = i + 1
        vals = [i, j['title'], j['status'], j['hours'], j['price'],
                j['apply_count'], j['publisher'], _clean_for_excel(j['description']), j['link']]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = THIN_BORDER
            if col == 1:
                c.alignment = Alignment(horizontal='center')

    widths = {'A': 6, 'B': 45, 'C': 18, 'D': 10, 'E': 12, 'F': 12, 'G': 18, 'H': 80, 'I': 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, len(jobs) + 2):
        ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)
    logger.info(f'[OK] 已保存: {path} ({len(jobs)}条)')


async def fetch_page(session, page_num: int, consecutive_403: list) -> tuple:
    """抓取单页，带403退避、失败重试、Referer链伪装"""
    url = BASE_URL if page_num == 1 else f'{BASE_URL}/page{page_num}'
    proxy_url = get_proxy_url()
    proxies = {'http': proxy_url, 'https': proxy_url} if proxy_url else {}

    # 构建 Referer：让请求看起来像从上一页或首页跳过来的
    referer = ''
    if page_num > 1:
        referer = f'https://www.yuanjisong.com/job/allcity/page{page_num - 1}'
    elif page_num == 1:
        referer = 'https://www.yuanjisong.com/'

    # 指数退避：连续 403 时等待更久
    if consecutive_403[0] > 0:
        wait = min(2 ** consecutive_403[0], 30)
        logger.warning(f'连续 {consecutive_403[0]} 次 403，等待 {wait}s 后重试第 {page_num} 页')
        await asyncio.sleep(wait)

    # 最多重试3次（超时/连接错误）
    for attempt in range(3):
        try:
            # 先访问首页建立 session/cookie
            home_resp = await session.get('https://www.yuanjisong.com/', timeout=10,
                                          impersonate='chrome131', allow_redirects=True, proxies=proxies)
            if home_resp.status_code != 200:
                logger.warning(f'第{page_num}页: 首页返回 {home_resp.status_code}，重试...')
                await asyncio.sleep(1)
                continue

            # 随机停顿（0.3~1.0s），避免固定间隔
            await asyncio.sleep(random.uniform(0.3, 1.0))

            # 带 Referer 请求目标页
            extra_headers = {'Referer': referer} if referer else {}
            resp = await session.get(url, impersonate='chrome131', timeout=15,
                                     allow_redirects=True, proxies=proxies,
                                     headers=extra_headers)
            if resp.status_code == 403:
                consecutive_403[0] += 1
                logger.warning(f'第{page_num}页: HTTP 403 (proxy={proxy_url or "直连"}) [{consecutive_403[0]}次连续]')
                if proxy_url:
                    await get_pool().mark_fail(proxy_url)
                return page_num, [], 0

            if resp.status_code != 200:
                logger.warning(f'第{page_num}页: HTTP {resp.status_code} (proxy={proxy_url or "直连"})')
                return page_num, [], 0

            # 重置 403 计数器（成功则说明恢复正常）
            consecutive_403[0] = 0

            jobs = parse_page(resp.text)
            all_cards = len(re.findall(r'<div class="job_card">', resp.text))

            # 记录代理成功
            if proxy_url:
                await get_pool().mark_success(proxy_url)

            # 每抓10页保存一次cookie
            if page_num % 10 == 0:
                save_cookies(session, COOKIE_FILE)
                logger.debug(f'第{page_num}页: 已保存Cookie')

            return page_num, jobs, all_cards

        except (cffi_requests.exceptions.Timeout, cffi_requests.exceptions.ConnectionError) as e:
            if attempt < 2:
                wait = min(2 ** (attempt + 1), 10)
                logger.warning(f'第{page_num}页超时/连接失败: {e}，{wait}s后重试 ({attempt+1}/3)')
                await asyncio.sleep(wait)
            else:
                logger.error(f'第{page_num}页最终失败: {e}')
                return page_num, [], 0
        except Exception as e:
            logger.error(f'第{page_num}页请求异常: {e}')
            return page_num, [], 0


def should_exclude(job: dict, keywords: list) -> bool:
    text = f"{job['title']} {job['description']}".lower()
    return any(kw.lower() in text for kw in keywords)


def parse_price(price_str: str) -> int:
    m = re.search(r'(\d+)', str(price_str))
    return int(m.group(1)) if m else 0


async def scrape(max_pages: int = 300, exclude: bool = True, resume: bool = False, output_dir: Path = None):
    if output_dir is None:
        output_dir = OUTPUT_DIR
    print('=' * 60)
    print('猿急送爬虫 - curl_cffi 轻量版 v3')
    print(f'时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)
    logger.info('启动 curl_cffi 爬虫...')

    # 初始化代理池（后台自动刷新）
    print('\n[代理池] 正在初始化...')
    pool_status = await init_pool()
    print(f'[代理池] 抓取 {pool_status["fetched"]} 个，可用 {pool_status["working"]} 个，耗时 {pool_status["elapsed"]}s')
    print()

    # 创建带 Cookie 的异步 session（复用连接）
    session = cffi_requests.AsyncSession(impersonate='chrome131')

    # 加载上次保存的 cookie（断点续爬时保持会话）
    if resume and load_cookies(session, COOKIE_FILE):
        logger.info('已恢复 Cookie，尝试断点续爬...')

    all_jobs = []
    seen_ids = set()
    # 断点续爬：从最后有效页继续
    start_page = load_last_page(LAST_PAGE_FILE) if resume else 1
    page_num = start_page
    empty_count = 0
    last_valid_page = start_page
    consecutive_403 = [0]

    if start_page > 1:
        print(f'[断点续爬] 从第 {start_page} 页继续')
        print()

    print(f'[目标] 共 {max_pages} 页，预计每页 3s，约 {max_pages * 3 // 60} 分钟')
    print()

    # ========== 主抓取循环 ==========
    while page_num <= max_pages and empty_count < 2:
        pn, jobs, total_cards = await fetch_page(session, page_num, consecutive_403)

        if total_cards == 0:
            empty_count += 1
            logger.warning(f'第{pn}页空页（连续{empty_count}）')
        else:
            empty_count = 0
            last_valid_page = pn
            save_last_page(pn, LAST_PAGE_FILE)  # 保存断点
            new_count = 0
            for j in jobs:
                if j['id'] in seen_ids:
                    continue
                seen_ids.add(j['id'])
                if exclude and should_exclude(j, EXCLUDE_KW):
                    continue
                all_jobs.append(j)
                new_count += 1
            eta = (max_pages - page_num) * 2.5  # 预估剩余秒数
            print(f'  第{pn}/{max_pages}页 | 累计{len(all_jobs)}条 | ETA {eta//60:.0f}分{eta%60:.0f}秒', flush=True)
            logger.success(f'第{pn}页: {total_cards}卡片, 新增{new_count}条, 累计{len(all_jobs)}')

        page_num += 1
        # 随机延迟：模拟人类浏览节奏（1.5~3.5秒）
        await asyncio.sleep(random.uniform(1.5, 3.5))

    # ========== 保存 ==========
    if all_jobs:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_excel(all_jobs, str(output_dir / f'yuanjisong_{ts}.xlsx'))

        # 学生项目（500元以内）
        student = [j for j in all_jobs if parse_price(j['price']) <= 500]
        if student:
            save_excel(student, str(output_dir / f'student_{ts}.xlsx'))
            logger.info(f'学生项目: {len(student)}条')

        # Python 相关
        python_kw = ['python', '爬虫', '脚本', '数据', '抓取', 'api', '接口', '自动化']
        python = [j for j in all_jobs
                  if any(kw in (j['title'] + j['description']).lower() for kw in python_kw)]
        if python:
            save_excel(python, str(output_dir / f'python_{ts}.xlsx'))
            logger.info(f'Python项目: {len(python)}条')
    else:
        logger.warning('未获取到任何可投递项目')

    print(f'\n[完成] 共 {len(all_jobs)} 条可投递项目，最后有效页: {last_valid_page}')
    print(f'[代理池] 最终状态: {get_pool().status()}')
    await session.close()
    await get_pool().stop()
    return all_jobs


def main():
    import argparse
    default_output = str(OUTPUT_DIR)
    parser = argparse.ArgumentParser(description='猿急送爬虫 - curl_cffi 轻量版 v3')
    parser.add_argument('--max-pages', type=int, default=300, help='最大页数（默认300）')
    parser.add_argument('--output-dir', type=str, default=default_output, help='输出目录')
    parser.add_argument('--include', action='store_true', help='不排除关键词（调试用）')
    parser.add_argument('--resume', action='store_true', help='断点续爬（从最后有效页继续）')
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    asyncio.run(scrape(
        max_pages=args.max_pages,
        exclude=not args.include,
        resume=args.resume,
        output_dir=output_dir,
    ))


if __name__ == '__main__':
    main()
