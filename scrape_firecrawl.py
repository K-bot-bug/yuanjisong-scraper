"""
猿急送爬虫 - Firecrawl 方案
==============================
用于绕过 IP 封禁，通过 Firecrawl 代理访问
使用方式：在 Claude Code 中调用 MCP 工具获取数据后运行此脚本解析保存

使用方法：
  方式1（推荐）：在 Claude Code 中手动调用 firecrawl_scrape 获取数据，再运行本脚本解析
  方式2：配置 FIRECRAWL_API_KEY 环境变量后直接运行

运行：
  python scrape_firecrawl.py --pages 10
  python scrape_firecrawl.py --pages 10 --include
"""
import asyncio
import json
import os
import re
import sys
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from loguru import logger

sys.stdout.reconfigure(encoding='utf-8')
logger.add("scrape_firecrawl.log", rotation="10 MB", encoding="utf-8")

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "数据")
os.makedirs(OUTPUT_DIR, exist_ok=True)

BASE_URL = "https://www.yuanjisong.com/job/allcity"

EXCLUDE_KW = [
    '架构', '高级', '专家', '深入', '底层', '内核', '编译原理', '逆向工程', '风控',
    '算法', '深度学习', '区块链', '量化', '挖矿', '渗透测试', '漏洞挖掘', '密码学',
    '安全研究', '反作弊', '反欺诈', '数字取证', '威胁情报', '性能优化',
    '接码平台', '翻墙', 'vpn代理', '绕过', '破解', '灰色', '套利', '刷单',
    '刷量', '黑产', '洗钱', '赌博', '彩票', '博彩', '棋牌', '网赌',
    '代写论文', '代做毕设', '学术不端',
    'stm32', '嵌入式', 'esp32', '树莓派', '单片机', 'iot', 'arduino',
    '传感器', '物联网', 'fpga', '电路板', '固件', 'arm开发',
    'app加固', '混淆', '加壳', 'so开发', '视频防嗅探',
    '游戏', 'unity', 'game', '网游', '手游',
    '相机', '标定', '镜头', '光学', '驻场',
]

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)


def parse_markdown(markdown_text: str) -> list[dict]:
    """从 Firecrawl 返回的 markdown 解析项目"""
    jobs = []
    # 匹配标题和链接
    title_links = re.findall(r'\*\*(.+?)\]\((https://www\.yuanjisong\.com/job/\d+)\)', markdown_text)

    for title, link in title_links:
        idx = markdown_text.find(f'**{title}**]({link})')
        if idx == -1:
            continue
        chunk = markdown_text[idx:idx + 2000]

        # 状态
        status_m = re.search(r'(项目制|时间制)\s*(.+?)(?:工时)', chunk)
        status = f'{status_m.group(1)} {status_m.group(2).strip()}' if status_m else ''

        # 工时
        hours_m = re.search(r'工时：(\d+\s*\w+)', chunk)
        hours = hours_m.group(1).strip() if hours_m else ''

        # 价格
        price_m = re.search(r'¥(\d+)_元_', chunk)
        price = f'¥{price_m.group(1)}元' if price_m else ''

        # 投递人数
        apply_m = re.search(r'\*\*(\d+)\*\*\s*人已投递', chunk)
        apply_count = f'{apply_m.group(1)}人已投递' if apply_m else ''

        # 发布者
        pub_m = re.search(r'\[(.+?)\]\(https://www\.yuanjisong\.com/employer/\d+\)', chunk)
        publisher = pub_m.group(1) if pub_m else ''

        # 描述
        desc_m = re.search(r'\[描述：(.+?)\]\(', chunk)
        description = desc_m.group(1).replace('\\\\n', '\n').replace('\\\\', '').strip() if desc_m else ''

        jobs.append({
            'id': link.split('/')[-1],
            'title': title,
            'status': status,
            'hours': hours,
            'price': price,
            'description': description,
            'apply_count': apply_count,
            'publisher': publisher,
            'link': link,
        })
    return jobs


def should_exclude(job: dict, keywords: list) -> bool:
    text = f"{job['title']} {job['description']}".lower()
    return any(kw.lower() in text for kw in keywords)


def parse_price(price_str: str) -> int:
    m = re.search(r'(\d+)', str(price_str))
    return int(m.group(1)) if m else 0


def save_excel(jobs: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = '可投递项目'

    headers = ['序号', '项目标题', '项目状态', '工时', '预算', '投递人数', '发布者', '项目描述', '项目链接']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = THIN_BORDER

    for i, j in enumerate(jobs, 1):
        row_idx = i + 1
        vals = [i, j['title'], j['status'], j['hours'], j['price'],
                j['apply_count'], j['publisher'], j['description'], j['link']]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = THIN_BORDER
            if col == 1:
                c.alignment = Alignment(horizontal='center')

    widths = {'A': 6, 'B': 45, 'C': 18, 'D': 10, 'E': 12, 'F': 12, 'G': 18, 'H': 80, 'I': 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(2, len(jobs) + 2):
        ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)
    logger.info(f'[OK] 已保存: {path} ({len(jobs)}条)')


async def scrape_from_markdown(markdown_pages: dict) -> list[dict]:
    """
    从 markdown 数据字典解析并保存
    markdown_pages: {page_num: markdown_text}
    """
    all_jobs = []
    seen_ids = set()

    for pn in sorted(markdown_pages.keys()):
        md = markdown_pages[pn]
        jobs = parse_markdown(md)
        new_count = 0
        for j in jobs:
            if j['id'] in seen_ids:
                continue
            seen_ids.add(j['id'])
            all_jobs.append(j)
            new_count += 1
        logger.success(f'第{pn}页: 解析到{len(jobs)}条, 新增{new_count}条, 累计{len(all_jobs)}')

    return all_jobs


def main():
    import argparse
    parser = argparse.ArgumentParser(description='猿急送爬虫 - Firecrawl 方案')
    parser.add_argument('--pages', type=int, default=10, help='抓取页数（默认10）')
    parser.add_argument('--include', action='store_true', help='不排除关键词')
    parser.add_argument('--input', type=str, help='输入 JSON 文件路径（包含 markdown 数据）')
    args = parser.parse_args()

    print('=' * 60)
    print('猿急送爬虫 - Firecrawl 方案')
    print(f'时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    print('=' * 60)

    if args.input:
        # 从文件加载
        with open(args.input, encoding='utf-8') as f:
            data = json.load(f)
        markdown_pages = data.get('pages', {})
        print(f'从文件加载 {len(markdown_pages)} 页数据')
    else:
        # 提示用户
        print(f'请提供 {args.pages} 页的 Firecrawl markdown 数据')
        print('在 Claude Code 中执行:')
        print(f'  循环调用 mcp__firecrawl__firecrawl_scrape(url="{BASE_URL}/page{{N}}") 获取每页 markdown')
        print()
        print('或者保存为一个 JSON 文件后运行:')
        print(f'  python scrape_firecrawl.py --input data.json --pages {args.pages}')
        return

    all_jobs = asyncio.run(scrape_from_markdown(markdown_pages))

    # 过滤
    if not args.include:
        suitable = [j for j in all_jobs if not should_exclude(j, EXCLUDE_KW)]
        print(f'过滤后: {len(suitable)} 条 (排除 {len(all_jobs) - len(suitable)} 条)')
    else:
        suitable = all_jobs

    # 保存
    if suitable:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_excel(suitable, os.path.join(OUTPUT_DIR, f'yuanjisong_{ts}.xlsx'))

        student = [j for j in suitable if parse_price(j['price']) <= 500]
        if student:
            save_excel(student, os.path.join(OUTPUT_DIR, f'student_{ts}.xlsx'))
            print(f'学生项目: {len(student)}条')

        python_kw = ['python', '爬虫', '脚本', '数据', '抓取', 'api', '接口', '自动化']
        python = [j for j in suitable
                  if any(kw in (j['title'] + j['description']).lower() for kw in python_kw)]
        if python:
            save_excel(python, os.path.join(OUTPUT_DIR, f'python_{ts}.xlsx'))
            print(f'Python项目: {len(python)}条')
    else:
        print('未解析到任何项目')

    print(f'\n[完成] 共 {len(suitable)} 条可投递项目')


if __name__ == '__main__':
    main()
